import openpyxl, os
os.chdir("C:\\Users\\User\\Desktop\\PROGRAMS\\python\\Pycharm Project Files\\documents\\")
w = openpyxl.Workbook()
sh = w["Sheet"]
for x in range(1,6):
    sh[f"A{x}"] = x
sh["B1"] = "mango"
sh["B2"] = "apple"
sh["B3"] = "orange"
sh["B4"] = "pineapple"
sh["B5"] = "strawberry"
w.save("excel.xlsx")
wb = openpyxl.load_workbook("excel.xlsx")
sheet = wb["Sheet"]
print(sheet["A1"].value)  # returns the value of the cell
print(sheet.cell(row=1, column=1).value) # returns the same as the before one
print("-"*100)
for i in range(1, 6):
    print(i, sheet.cell(row=i, column=2).value)
wb.close()